from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django.contrib import messages
from .models import Machine, Sparepart, JadwalPreventive
from .forms import MachineForm, SparepartForm, JadwalForm
from django.http import HttpResponse
import openpyxl
from datetime import datetime
import calendar
from collections import defaultdict
from django.db.models import Q

# --- DASHBOARD ---
def dashboard(request):
    # 1. DATA STATISTIK
    total_spareparts = Sparepart.objects.count()
    total_machines = Machine.objects.count()
    low_stock_items = Sparepart.objects.filter(stock__lte=models.F('min_stock'))
    machines = Machine.objects.all().order_by('code') # Urutkan sesuai kode
    
    # Jadwal Pending (Untuk Small Box)
    jadwal_preventive = JadwalPreventive.objects.filter(status='Pending').order_by('tgl_jadwal')

    # 2. DATA KALENDER TAHUNAN
    year = datetime.now().year
    
    # Ambil semua mesin dan jadwal di tahun ini
    all_events = JadwalPreventive.objects.filter(tgl_jadwal__year=year).select_related('machine')
    
    # Kelompokkan Jadwal per Mesin
    events_by_machine = defaultdict(list)
    for e in all_events:
        events_by_machine[e.machine.id].append(e)

    # Siapkan Data Baris untuk Tabel
    rows = []
    bulan_indo = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    
    for m in machines:
        month_events = [[] for _ in range(12)]
        if m.id in events_by_machine:
            for e in events_by_machine[m.id]:
                month_events[e.tgl_jadwal.month - 1].append(e)
        
        rows.append({
            'machine': m,
            'events': month_events
        })
    
    # 3. GABUNGKAN KE CONTEXT
    context = {
        'total_spareparts': total_spareparts,
        'total_machines': total_machines,
        'low_stock_items': low_stock_items,
        'machines': machines,
        'jadwal_preventive': jadwal_preventive,
        'year': year,
        'months_name': bulan_indo[1:],
        'rows': rows,
    }
    return render(request, 'dashboard.html', context)

# --- SPAREPART ---
def sparepart_list(request):
    query = request.GET.get('q')
    machine_id = request.GET.get('machine')

    machines = Machine.objects.all()
    selected_machine = None

    # Mulai ambil semua data sparepart
    spareparts = Sparepart.objects.all()

    # --- LOGIKA SEARCH & FILTER (SUDAH DIPERBAIKI) ---
    if query:
        # Pencarian berdasarkan: Nama Barang, Kode Barang, Nama Mesin, Kode Mesin
        spareparts = spareparts.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(machine__name__icontains=query) |
            Q(machine__code__icontains=query)
        )
    elif machine_id:
        # Filter berdasarkan mesin (dari sidebar)
        spareparts = Sparepart.objects.filter(machine_id=machine_id)
        selected_machine = Machine.objects.get(id=machine_id)
    
    # --- HANDLING FORM TAMBAH DATA (POST) ---
    if request.method == 'POST':
        form = SparepartForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data sparepart berhasil disimpan.')
            return redirect('sparepart_list')
    else:
        form = SparepartForm()
    
    return render(request, 'sparepart_list.html', {
        'spareparts': spareparts, 
        'form': form, 
        'machines': machines,
        'selected_machine': selected_machine
    })

def sparepart_edit(request, pk):
    sparepart = get_object_or_404(Sparepart, pk=pk)
    if request.method == 'POST':
        form = SparepartForm(request.POST, request.FILES, instance=sparepart)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data berhasil diperbarui.')
            return redirect('sparepart_list')
    else:
        form = SparepartForm(instance=sparepart)
        
    return render(request, 'sparepart_form.html', {'form': form})

def sparepart_delete(request, pk):
    obj = get_object_or_404(Sparepart, pk=pk)
    obj.delete()
    messages.success(request, 'Data dihapus.')
    return redirect('sparepart_list')

# --- EXPORT EXCEL SPAREPART ---
def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Stok"

    headers = ['Kode Barang', 'Nama Sparepart', 'Kode Mesin', 'Nama Mesin', 'Lokasi Penyimpanan', 'Stok']
    ws.append(headers)

    items = Sparepart.objects.all()

    for item in items:
        ws.append([
            item.code,                
            item.name,                
            item.machine.code,        
            item.machine.name,        
            item.lokasi_penyimpanan,  
            item.stock                
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Stok_Meidoh.xlsx'
    wb.save(response)
    return response

# --- MESIN ---
def machine_list(request):
    machines = Machine.objects.all()
    if request.method == 'POST':
        form = MachineForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mesin berhasil ditambahkan.')
            return redirect('machine_list')
    else:
        form = MachineForm()

    return render(request, 'machine_list.html', {'machines': machines, 'form': form})

# --- JADWAL PREVENTIVE ---
def jadwal_list(request):
    query = request.GET.get('q')
    machine_id = request.GET.get('machine')

    machines = Machine.objects.all()
    selected_machine = None

    # Filter jadwal
    jadwal_list = JadwalPreventive.objects.all().order_by('-tgl_jadwal')
    
    if query:
        jadwal_list = jadwal_list.filter(Q(keterangan__icontains=query) | Q(machine__name__icontains=query))
    elif machine_id: 
        jadwal_list = jadwal_list.filter(machine_id=machine_id)
        selected_machine = Machine.objects.get(id=machine_id)
        
    # Hitung ringkasan
    total_jadwal = jadwal_list.count()
    pending_count = jadwal_list.filter(status='Pending').count()
    done_count = jadwal_list.filter(status='Selesai').count()
    
    context = {
        'jadwal_list': jadwal_list,
        'total_jadwal': total_jadwal,
        'pending_count': pending_count,
        'done_count': done_count,
    }
    return render(request, 'jadwal_list.html', context)

def tambah_jadwal(request):
    if request.method == 'POST':
        form = JadwalForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('jadwal_list')
    else:
        form = JadwalForm()
    
    return render(request, 'jadwal_form.html', {'form': form})

def jadwal_edit(request, pk):
    jadwal = get_object_or_404(JadwalPreventive, pk=pk)
    if request.method == 'POST':
        form = JadwalForm(request.POST, instance=jadwal)
        if form.is_valid():
            form.save()
            messages.success(request, 'Jadwal berhasil diperbarui.')
            return redirect('jadwal_list')
    else:
        form = JadwalForm(instance=jadwal)
        
    return render(request, 'jadwal_form.html', {'form': form})

def jadwal_delete(request, pk):
    obj = get_object_or_404(JadwalPreventive, pk=pk)
    obj.delete()
    messages.success(request, 'Jadwal dihapus.')
    return redirect('jadwal_list')

# --- KALENDER MAINTENANCE ---
def jadwal_kalender(request):
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))

    # Navigasi Bulan
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
        
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    cal = calendar.Calendar()
    weeks = cal.monthdayscalendar(year, month)

    jadwal_list = JadwalPreventive.objects.filter(
        tgl_jadwal__year=year, 
        tgl_jadwal__month=month
    )

    jadwal_by_day = defaultdict(list)
    for item in jadwal_list:
        jadwal_by_day[item.tgl_jadwal.day].append(item)

    bulan_indo = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    month_name = bulan_indo[month]

    context = {
        'year': year,
        'month': month,
        'month_name': month_name,
        'weeks': weeks,
        'jadwal_list': jadwal_list,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
    }
    return render(request, 'jadwal_kalender.html', context)

def jadwal_kalender_tahunan(request):
    year = int(request.GET.get('year', datetime.now().year))
    prev_year = year - 1
    next_year = year + 1

    # Ambil Semua Mesin
    machines = Machine.objects.all()

    # Ambil SemUA Jadwal di Tahun Tersebut
    events = JadwalPreventive.objects.filter(tgl_jadwal__year=year).select_related('machine')

    # Kelompokkan Jadwal berdasarkan ID Mesin
    events_by_machine = defaultdict(list)
    for e in events:
        events_by_machine[e.machine.id].append(e)

    # Siapkan Data Baris (Rows)
    rows = []
    bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    
    for m in machines:
        # Siapkan array kosong untuk 12 bulan
        month_events = [[] for _ in range(12)]
        
        # Jika mesin ini punya jadwal, masukkan ke array bulan yang sesuai
        if m.id in events_by_machine:
            for e in events_by_machine[m.id]:
                month_index = e.tgl_jadwal.month - 1 # (Bulan 1 = Index 0)
                month_events[month_index].append(e)
        
        rows.append({
            'machine': m,
            'events': month_events
        })

    context = {
        'year': year,
        'prev_year': prev_year,
        'next_year': next_year,
        'months_name': bulan_indo,
        'rows': rows,
    }
    return render(request, 'jadwal_kalender_tahunan.html', context)

def export_excel_tahunan(request):
    year = request.GET.get('year', datetime.now().year)
    
    # Ambil data tahun tersebut
    items = JadwalPreventive.objects.filter(tgl_jadwal__year=year).select_related('machine')

    # Buat Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal_Maintenance_{year}"

    # Header
    headers = ['No', 'Tanggal', 'Nama Mesin', 'Kode Mesin', 'Jenis Maintenance', 'Status', 'Teknisi', 'Keterangan']
    ws.append(headers)

    # Styling Header
    for col in ws[1]:
        col.font = openpyxl.styles.Font(bold=True)
        col.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Isi Data
    for index, item in enumerate(items, start=1):
        ws.append([
            index,
            item.tgl_jadwal,
            item.machine.name,
            item.machine.code,
            item.jenis_maintenance,
            item.status,
            item.teknisi,
            item.keterangan
        ])

    # Kolom Auto Width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Maintenance_Tahunan_{year}.xlsx'
    wb.save(response)
    return response