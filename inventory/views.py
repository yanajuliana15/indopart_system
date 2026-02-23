from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django.contrib import messages
from .models import Machine, Sparepart, JadwalPreventive, Breakdown
from .forms import MachineForm, SparepartForm, JadwalForm, BreakdownForm
from django.http import HttpResponse
import openpyxl
from datetime import datetime
import calendar
from collections import defaultdict
from django.db.models import Q

# --- DASHBOARD ---
def dashboard(request):
    # 1. DATA STATISTIK
    total_spareparts = Sparepart.objects.count()
    total_machines = Machine.objects.count()
    total_breakdowns = Breakdown.objects.count()
    
    low_stock_items = Sparepart.objects.filter(stock__lte=models.F('min_stock'))
    machines = Machine.objects.all().order_by('code') 
    jadwal_preventive = JadwalPreventive.objects.filter(status='Pending').order_by('tgl_jadwal')
    recent_breakdowns = Breakdown.objects.all().order_by('-created_at')[:5]

    # 2. DATA KALENDER TAHUNAN
    year = datetime.now().year
    all_events = JadwalPreventive.objects.filter(tgl_jadwal__year=year).select_related('machine')
    
    events_by_machine = defaultdict(list)
    for e in all_events:
        events_by_machine[e.machine.id].append(e)

    bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    
    rows = []
    for month_idx in range(12):
        month_name = bulan_indo[month_idx]
        machine_events_in_month = []
        for m in machines:
            evs = []
            if m.id in events_by_machine:
                for e in events_by_machine[m.id]:
                    if e.tgl_jadwal.month - 1 == month_idx:
                        evs.append(e)
            machine_events_in_month.append(evs)
        
        rows.append({
            'month': month_name,
            'events': machine_events_in_month
        })

    context = {
        'total_spareparts': total_spareparts,
        'total_machines': total_machines,
        'total_breakdowns': total_breakdowns,
        'low_stock_items': low_stock_items,
        'machines': machines,
        'jadwal_preventive': jadwal_preventive,
        'recent_breakdowns': recent_breakdowns,
        'year': year,
        'months_name': bulan_indo,
        'rows': rows,
    }
    return render(request, 'dashboard.html', context)

# --- SPAREPART ---
def sparepart_list(request):
    query = request.GET.get('q')
    machine_id = request.GET.get('machine')
    machines = Machine.objects.all()
    selected_machine = None
    spareparts = Sparepart.objects.all()

    if query:
        spareparts = spareparts.filter(
            Q(name__icontains=query) | Q(code__icontains=query) | 
            Q(machine__name__icontains=query) | Q(machine__code__icontains=query)
        )
    elif machine_id:
        spareparts = Sparepart.objects.filter(machine_id=machine_id)
        selected_machine = Machine.objects.get(id=machine_id)
    
    if request.method == 'POST':
        form = SparepartForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data sparepart berhasil disimpan.')
            return redirect('sparepart_list')
    else:
        form = SparepartForm()
    
    return render(request, 'sparepart_list.html', {
        'spareparts': spareparts, 'form': form, 'machines': machines,
        'selected_machine': selected_machine
    })

def sparepart_edit(request, pk):
    sparepart = get_object_or_404(Sparepart, pk=pk)
    if request.method == 'POST':
        form = SparepartForm(request.POST, request.FILES, instance=sparepart)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data berhasil diperbarui.')
            return redirect('sparepart_list')
    else:
        form = SparepartForm(instance=sparepart)
    return render(request, 'sparepart_form.html', {'form': form})

def sparepart_delete(request, pk):
    obj = get_object_or_404(Sparepart, pk=pk)
    obj.delete()
    messages.success(request, 'Data dihapus.')
    return redirect('sparepart_list')

def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Stok"
    headers = ['Kode Barang', 'Nama Sparepart', 'Kode Mesin', 'Nama Mesin', 'Lokasi Penyimpanan', 'Stok']
    ws.append(headers)
    items = Sparepart.objects.all()
    for item in items:
        ws.append([item.code, item.name, item.machine.code, item.machine.name, item.lokasi_penyimpanan, item.stock])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Stok_Meidoh.xlsx'
    wb.save(response)
    return response

# --- MESIN ---
def machine_list(request):
    machines = Machine.objects.all().order_by('name') 
    if request.method == 'POST':
        form = MachineForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mesin berhasil ditambahkan.')
            return redirect('machine_list')
    else:
        form = MachineForm()
    return render(request, 'machine_list.html', {'machines': machines, 'form': form})

def machine_edit(request, pk):
    # Indentasi diperbaiki: Fungsi ini sekarang berada di level global, bukan di dalam fungsi lain
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        form = MachineForm(request.POST, request.FILES, instance=machine)
        if form.is_valid():
            form.save()
            return redirect('machine_list')
    else:
        form = MachineForm(instance=machine)
    return render(request, 'machine_form.html', {'form': form, 'machine': machine})

def machine_delete(request, pk):
    # Indentasi diperbaiki
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        machine.delete()
        return redirect('machine_list')
    return redirect('machine_list')

# --- JADWAL PREVENTIVE ---
def jadwal_list(request):
    query = request.GET.get('q')
    machine_id = request.GET.get('machine')
    machines = Machine.objects.all()
    selected_machine = None
    jadwal_list = JadwalPreventive.objects.all().order_by('-tgl_jadwal')
    
    if query:
        jadwal_list = jadwal_list.filter(Q(keterangan__icontains=query) | Q(machine__name__icontains=query))
    elif machine_id: 
        jadwal_list = jadwal_list.filter(machine_id=machine_id)
        selected_machine = Machine.objects.get(id=machine_id)
        
    total_jadwal = jadwal_list.count()
    pending_count = jadwal_list.filter(status='Pending').count()
    done_count = jadwal_list.filter(status='Selesai').count()
    
    context = {
        'jadwal_list': jadwal_list, 'total_jadwal': total_jadwal,
        'pending_count': pending_count, 'done_count': done_count,
    }
    return render(request, 'jadwal_list.html', context)

def tambah_jadwal(request):
    if request.method == 'POST':
        form = JadwalForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('jadwal_list')
    else:
        form = JadwalForm()
    return render(request, 'jadwal_form.html', {'form': form})

def jadwal_edit(request, pk):
    jadwal = get_object_or_404(JadwalPreventive, pk=pk)
    if request.method == 'POST':
        form = JadwalForm(request.POST, instance=jadwal)
        if form.is_valid():
            form.save()
            messages.success(request, 'Jadwal berhasil diperbarui.')
            return redirect('jadwal_list')
    else:
        form = JadwalForm(instance=jadwal)
    return render(request, 'jadwal_form.html', {'form': form})

def jadwal_delete(request, pk):
    obj = get_object_or_404(JadwalPreventive, pk=pk)
    obj.delete()
    messages.success(request, 'Jadwal dihapus.')
    return redirect('jadwal_list')

# --- KALENDER MAINTENANCE ---
def jadwal_kalender(request):
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    if month == 12: next_year, next_month = year + 1, 1
    else: next_year, next_month = year, month + 1
    if month == 1: prev_year, prev_month = year - 1, 12
    else: prev_year, prev_month = year, month - 1

    cal = calendar.Calendar()
    weeks = cal.monthdayscalendar(year, month)
    jadwal_list = JadwalPreventive.objects.filter(tgl_jadwal__year=year, tgl_jadwal__month=month)
    jadwal_by_day = defaultdict(list)
    for item in jadwal_list: jadwal_by_day[item.tgl_jadwal.day].append(item)

    bulan_indo = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    month_name = bulan_indo[month]
    context = {'year': year, 'month': month, 'month_name': month_name, 'weeks': weeks, 'jadwal_list': jadwal_list, 'prev_year': prev_year, 'prev_month': prev_month, 'next_year': next_year, 'next_month': next_month}
    return render(request, 'jadwal_kalender.html', context)

def jadwal_kalender_tahunan(request):
    year = int(request.GET.get('year', datetime.now().year))
    prev_year = year - 1
    next_year = year + 1
    
    machines = Machine.objects.all().order_by('name')
    events = JadwalPreventive.objects.filter(tgl_jadwal__year=year).select_related('machine')
    
    events_by_machine = defaultdict(list)
    for e in events:
        events_by_machine[e.machine.id].append(e)
    
    rows = []
    bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    
    for month_idx in range(12):
        month_name = bulan_indo[month_idx]
        machine_events_in_month = []
        for m in machines:
            evs = []
            if m.id in events_by_machine:
                for e in events_by_machine[m.id]:
                    if e.tgl_jadwal.month - 1 == month_idx:
                        evs.append(e)
            machine_events_in_month.append(evs)
        
        rows.append({
            'month': month_name,
            'events': machine_events_in_month
        })

    context = {
        'year': year, 'prev_year': prev_year, 'next_year': next_year,
        'months_name': bulan_indo, 
        'rows': rows, 
        'machines': machines
    }
    return render(request, 'jadwal_kalender_tahunan.html', context)

def export_excel_tahunan(request):
    year = request.GET.get('year', datetime.now().year)
    items = JadwalPreventive.objects.filter(tgl_jadwal__year=year).select_related('machine')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal_Maintenance_{year}"
    headers = ['No', 'Tanggal', 'Nama Mesin', 'Kode Mesin', 'Jenis Maintenance', 'Status', 'Teknisi', 'Keterangan']
    ws.append(headers)
    for col in ws[1]:
        col.font = openpyxl.styles.Font(bold=True)
        col.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for index, item in enumerate(items, start=1):
        ws.append([index, item.tgl_jadwal, item.machine.name, item.machine.code, item.jenis_maintenance, item.status, item.teknisi, item.keterangan])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Maintenance_Tahunan_{year}.xlsx'
    wb.save(response)
    return response

# --- BREAKDOWN ---
# HAPUS DEFINISI FUNGSI YANG DUPLIKAT, GUNAKAN YANG LENGKAP INI
def breakdown_list(request):
    query = request.GET.get('q')
    machine_id = request.GET.get('machine')
    machines = Machine.objects.all()
    selected_machine = None
    breakdowns = Breakdown.objects.all().order_by('-tanggal')

    if query:
        breakdowns = breakdowns.filter(
            Q(penyebab__icontains=query) | Q(machine__name__icontains=query) | Q(pic__icontains=query) | Q(kerusakan__icontains=query) 
        )
    elif machine_id:
        breakdowns = breakdowns.filter(machine_id=machine_id)
        selected_machine = Machine.objects.get(id=machine_id)

    total_breakdown = breakdowns.count()
    context = {
        'breakdowns': breakdowns, 
        'machines': machines, 
        'selected_machine': selected_machine, 
        'total_breakdown': total_breakdown
    }
    return render(request, 'breakdown_list.html', context)

def tambah_breakdown(request):
    if request.method == 'POST':
        form = BreakdownForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Laporan breakdown dan foto berhasil disimpan.')
            return redirect('breakdown_list')
    else:
        form = BreakdownForm()
    return render(request, 'breakdown_form.html', {'form': form})

def breakdown_edit(request, pk):
    breakdown = get_object_or_404(Breakdown, pk=pk)
    if request.method == 'POST':
        form = BreakdownForm(request.POST, request.FILES, instance=breakdown)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data breakdown dan foto berhasil diperbarui.')
            return redirect('breakdown_list')
    else:
        form = BreakdownForm(instance=breakdown)
    return render(request, 'breakdown_form.html', {'form': form})

def breakdown_delete(request, pk):
    obj = get_object_or_404(Breakdown, pk=pk)
    obj.delete()
    messages.success(request, 'Data breakdown dihapus.')
    return redirect('breakdown_list')

def export_excel_breakdown(request):
    items = Breakdown.objects.all().select_related('machine', 'sparepart').order_by('-tanggal')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Breakdown"

    # Header
    headers = ['No', 'Tanggal', 'Nama Mesin', 'Kode Mesin', 'Kerusakan','Jenis Trouble', 'Penyebab', 'Tindakan', 'Sparepart Dipakai', 'PIC']
    ws.append(headers)

    # Style Header
    for col in ws[1]:
        col.font = openpyxl.styles.Font(bold=True)
        col.fill = openpyxl.styles.PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")

    # Isi Data
    for index, item in enumerate(items, start=1):
        sparepart_name = item.sparepart.name if item.sparepart else "-"
        
        # PERBAIKAN URUTAN: Sesuaikan dengan urutan Header
        ws.append([
            index,
            item.tanggal,
            item.machine.name,
            item.machine.code,      # Pindah ke sini (sebelumnya setelah kerusakan)
            item.kerusakan,         # Pindah ke sini (sebelumnya sebelum code)
            item.jenis_trouble,
            item.penyebab,
            item.tindakan,
            sparepart_name,
            item.pic
        ])

    # Auto Width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Laporan_Breakdown.xlsx'
    wb.save(response)
    return response